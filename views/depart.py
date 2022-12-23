import os
from openpyxl import load_workbook

from django.shortcuts import render, redirect, HttpResponse

from app01 import models
from app01.utils.pagination import Pagination


def depart_list(request):
    """部门列表"""
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加页面"""
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """编辑"""
    if request.method == 'GET':
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {'row_object': row_object})
    title = request.POST.get("title")  # 获取用户提交的标题
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


def depart_multi(request):
    """批量上传（基于excel）"""
    file_object = request.FILES.get("exc")  # 获取文件对象
    wb = load_workbook(file_object)         # 读取文件对象
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart/list/")

